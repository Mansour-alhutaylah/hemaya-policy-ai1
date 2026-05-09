"""Text extraction helpers for uploaded policy files.

Two parallel APIs:

  extract_text(file_path, file_ext) -> str
      The original API. Joined plain text. Used by every existing caller
      (chat, RAG engine, framework loader, legacy code paths). Signature
      and return shape are preserved verbatim for backwards compatibility.

  extract_text_segments(file_path, file_ext) -> list[dict]
      Phase 11. Returns the same content as a list of source-aware
      segments so downstream chunking can attach a page_number (PDF) or
      paragraph_index (DOCX) to each chunk. Used only by the upload
      pipeline and the opt-in rechunk path.

  Each segment is one of:
      {"text": str, "page_number": int, "paragraph_index": None}   # PDF
      {"text": str, "page_number": None, "paragraph_index": int}   # DOCX
      {"text": str, "page_number": None, "paragraph_index": None}  # TXT/XLSX

extract_text() is implemented in terms of extract_text_segments() — single
source of truth for the extraction loop, no duplication.
"""
from pathlib import Path


def extract_text(file_path: Path, file_ext: str) -> str:
    """Plain joined text. Backwards-compatible with all pre-Phase-11 callers."""
    try:
        segments = extract_text_segments(file_path, file_ext)
    except Exception as e:
        return f"[Extraction error: {e}]"
    return "\n".join(s["text"] for s in segments if s.get("text"))


def extract_text_segments(file_path: Path, file_ext: str) -> list:
    """Source-aware segmented extraction. Phase 11.

    Returns a list of {"text", "page_number", "paragraph_index"} dicts.
    Empty list on unsupported file types or extraction errors.
    """
    ext = file_ext.lower().lstrip(".")
    if ext == "pdf":
        return _extract_pdf_segments(file_path)
    if ext == "docx":
        return _extract_docx_segments(file_path)
    if ext in ("xlsx", "xls"):
        return _extract_xlsx_segments(file_path)
    if ext == "txt":
        return _extract_txt_segments(file_path)
    return []


def _extract_pdf_segments(file_path: Path) -> list:
    """One segment per page; page_number is 1-indexed for human display."""
    import fitz  # PyMuPDF
    doc = fitz.open(str(file_path))
    segments = []
    for page in doc:
        try:
            # sort=True asks PyMuPDF to return blocks in reading order,
            # which also corrects Arabic RTL bidi ordering and preserves
            # table-cell boundaries that get lost in flat get_text() mode.
            page_dict = page.get_text("dict", sort=True)
            text = _extract_page_blocks(page_dict)
        except Exception:
            # Per-page fallback: if structured extraction fails for any
            # reason (unusual PDF, corrupt page), degrade to plain text
            # for that page only — the rest of the document is unaffected.
            text = page.get_text()
        if text:
            segments.append({
                "text": text,
                "page_number": page.number + 1,  # 1-indexed for human display
                "paragraph_index": None,
            })
    return segments


def _extract_page_blocks(page_dict: dict) -> str:
    """Reconstruct page text from PyMuPDF's block/line/span structure.

    Type-0 blocks are text; type-1 are images (skipped).
    Lines within each block are joined with newlines;
    blocks are separated by blank lines to preserve paragraph boundaries.
    """
    blocks_text = []
    for block in page_dict.get("blocks", []):
        if block.get("type") != 0:  # skip image blocks
            continue
        line_texts = []
        for line in block.get("lines", []):
            span_text = " ".join(
                span["text"]
                for span in line.get("spans", [])
                if span.get("text", "").strip()
            )
            if span_text.strip():
                line_texts.append(span_text.strip())
        if line_texts:
            blocks_text.append("\n".join(line_texts))
    return "\n\n".join(blocks_text)


def _extract_docx_segments(file_path: Path) -> list:
    """One segment per non-empty paragraph; paragraph_index is 0-indexed enumerate."""
    from docx import Document
    doc = Document(str(file_path))
    segments = []
    for i, p in enumerate(doc.paragraphs):
        if p.text.strip():
            segments.append({
                "text": p.text,
                "page_number": None,
                "paragraph_index": i,
            })
    return segments


def _extract_xlsx_segments(file_path: Path) -> list:
    """Single segment with no source attribution (XLSX has no page concept)."""
    from openpyxl import load_workbook
    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = "\t".join(str(c) if c is not None else "" for c in row)
            if row_text.strip():
                lines.append(row_text)
    text = "\n".join(lines)
    if not text:
        return []
    return [{"text": text, "page_number": None, "paragraph_index": None}]


def _extract_txt_segments(file_path: Path) -> list:
    """Single segment with no source attribution (TXT has no page concept)."""
    import chardet
    raw = file_path.read_bytes()
    detected = chardet.detect(raw)
    encoding = detected.get("encoding") or "utf-8"
    text = raw.decode(encoding, errors="replace")
    if not text:
        return []
    return [{"text": text, "page_number": None, "paragraph_index": None}]
